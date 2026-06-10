"""
exporter.py — Fill ANY supplier/invoice Excel template by reading its headers.

Instead of a hard-coded column layout, the exporter scans the uploaded
template's header row(s), matches each header (English or Arabic) to an
extracted field, and writes the data into whatever columns that template
uses. So a template with extra columns (Sort code, Swift Code, …) or a
different column order just works — each matching column gets filled.

Falls back to the legacy fixed UAE layout only if no headers can be matched.
"""

import copy
import csv
import io
import os
import re
import tempfile
import warnings

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.datavalidation import DataValidation

from validator import canonical_emirate

SHEET_NAME = "Supplier Template"
START_ROW = 3                       # legacy fallback data row

# Legacy fixed map — only used if header auto-detection finds nothing.
COLUMN_MAP = [
    (2, "arabic_name"), (3, "english_name"), (7, "emirate"),
    (11, "bank_name_ar"), (13, "account_number"), (14, "iban"), (15, "english_name"),
]

# UAE-template dropdowns openpyxl drops on save (re-attached only for that sheet).
LIST_VALIDATIONS = [
    ("Sheet1!$B$2:$B$14", "G3:G251"),   # Emirate
    ("Sheet1!$A$2:$A$3",  "I3:I665"),   # Registration
]

CSV_FIELDS = [
    "source_file", "english_name", "arabic_name", "iban", "account_number",
    "swift", "currency", "bank_name_ar", "bank_name_en", "nationality",
    "emirates_id", "date_of_birth", "id_expiry", "emirate", "iban_valid",
]

# ── header → field synonyms (English + Arabic). Order/length disambiguates. ──
FIELD_SYNONYMS = {
    "arabic_name":   ["arabic vendor name", "arabic name", "name in arabic",
                      "vendor name arabic", "اسم المورد باللغة العربية",
                      "الاسم بالعربية", "اسم المورد"],
    "english_name":  ["english vendor name", "english name", "name in english",
                      "account name", "account title", "account holder",
                      "beneficiary name", "vendor name", "supplier name",
                      "company name", "اسم المورد باللغة الانجليزية", "اسم الحساب البنكي",
                      "اسم الحساب"],
    "iban":          ["iban number", "iban no", "iban",
                      "international bank account number",
                      "رقم الحساب المصرفي الدولي", "رقم الايبان", "ايبان"],
    "account_number":["account number", "bank account number", "account no", "acc no",
                      "رقم الحساب البنكي", "رقم الحساب المصرفي", "رقم الحساب"],
    "swift":         ["swift code", "swift bic", "swift", "bic code", "bic",
                      "رمز السويفت", "سويفت"],
    "sort_code":     ["sort code", "sortcode", "رمز الفرز"],
    "branch":        ["branch name", "branch", "اسم الفرع", "الفرع"],
    "bank_name":     ["bank name", "name of bank", "bank", "اسم البنك"],
    "emirate":       ["emirate", "emirates", "الامارة", "الإمارة"],
    "nationality":   ["nationality", "الجنسية"],
    "emirates_id":   ["emirates id number", "emirates id", "id number", "eid", "رقم الهوية"],
    "currency":      ["currency", "ccy", "العملة"],
    "date_of_birth": ["date of birth", "birth date", "dob", "تاريخ الميلاد"],
    "id_expiry":     ["id expiry", "expiry date", "expiry", "تاريخ الانتهاء"],
}

ARABIC_RE = re.compile(r"[؀-ۿ]")
TEXT_FIELDS = {"account_number", "iban", "emirates_id"}     # keep as literal text


def _norm(v) -> str:
    """Lowercase, unify Arabic alef/hamza, drop punctuation, collapse spaces."""
    s = str(v if v is not None else "").lower()
    s = s.replace("أ", "ا").replace("إ", "ا").replace("آ", "ا")   # unify alef forms
    s = re.sub(r"[^\w؀-ۿ]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


# Pre-build exact + substring lookups (substring sorted longest-first).
_EXACT, _SUBSTR = {}, []
for _field, _syns in FIELD_SYNONYMS.items():
    for _syn in _syns:
        _ns = _norm(_syn)
        _EXACT.setdefault(_ns, _field)
        _SUBSTR.append((_ns, _field))
_SUBSTR.sort(key=lambda x: -len(x[0]))


def _match_field(header) -> str:
    """Map a header string to a field key (or '' if none matches)."""
    h = _norm(header)
    if not h:
        return ""
    if h in _EXACT:
        return _EXACT[h]
    for ns, field in _SUBSTR:
        if ns and ns in h:
            return field
    return ""


def _value_for(field: str, rec: dict, header_text: str) -> str:
    """Resolve the cell value for a mapped field."""
    if field == "bank_name":
        if ARABIC_RE.search(header_text or ""):
            return rec.get("bank_name_ar") or rec.get("bank_name_en") or ""
        return rec.get("bank_name_en") or rec.get("bank_name_ar") or ""
    if field == "emirate":
        return canonical_emirate(rec.get("emirate", ""))
    return rec.get(field, "") or ""


def _detect_layout(ws, max_scan: int = 10):
    """Find the header row(s) and build {col: (field, combined_header)}.

    Returns dict(mappings=[(col, field, header_text)], data_start, header_rows)
    or None if no headers could be matched.
    """
    last = min(ws.max_row, max_scan)
    counts = {r: sum(1 for c in range(1, ws.max_column + 1)
                     if _match_field(ws.cell(r, c).value))
              for r in range(1, last + 1)}
    if not counts or max(counts.values()) == 0:
        return None

    best_n = max(counts.values())
    best_row = min(r for r in counts if counts[r] == best_n)     # topmost on tie
    threshold = max(2, best_n // 2)

    # Expand to the full contiguous block of header-like rows (above AND below),
    # so bilingual templates (English row + Arabic row) keep both headers.
    top = best_row
    while top - 1 >= 1 and counts.get(top - 1, 0) >= threshold:
        top -= 1
    bottom = best_row
    while bottom + 1 <= last and counts.get(bottom + 1, 0) >= threshold:
        bottom += 1
    header_rows = list(range(top, bottom + 1))

    mappings = []
    for c in range(1, ws.max_column + 1):
        combined = " ".join(
            str(ws.cell(hr, c).value) for hr in header_rows if ws.cell(hr, c).value
        )
        field = _match_field(combined)
        if field:
            mappings.append((c, field, combined))

    key_cols = [c for c, f, _ in mappings
                if f in ("english_name", "arabic_name", "iban", "account_number")] \
        or [c for c, _, _ in mappings]

    data_start = header_rows[-1] + 1
    while data_start <= ws.max_row and any(
        ws.cell(data_start, c).value not in (None, "") for c in key_cols
    ):
        data_start += 1
    return {"mappings": mappings, "data_start": data_start, "header_rows": header_rows}


def _pick_sheet(wb):
    """Choose the data sheet — the named template, else the best-matching sheet."""
    if SHEET_NAME in wb.sheetnames:
        return wb[SHEET_NAME]
    best, best_n = wb.active, -1
    for ws in wb.worksheets:
        layout = _detect_layout(ws)
        n = len(layout["mappings"]) if layout else 0
        if n > best_n:
            best, best_n = ws, n
    return best


def _right_align(cell):
    a = cell.alignment
    cell.alignment = Alignment(horizontal="right", vertical=a.vertical, readingOrder=2)


def _reattach_dropdowns(wb, ws):
    """Re-add list validations openpyxl drops — only for the UAE template."""
    if ws.title != SHEET_NAME or "Sheet1" not in wb.sheetnames:
        return
    try:
        for formula, cells in LIST_VALIDATIONS:
            dv = DataValidation(type="list", formula1=formula, allow_blank=True)
            dv.add(cells)
            ws.add_data_validation(dv)
    except Exception:
        pass


def existing_record_count(template_bytes: bytes) -> int:
    """How many records the template already holds (after its header rows)."""
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            wb = load_workbook(io.BytesIO(template_bytes))
        ws = _pick_sheet(wb)
        layout = _detect_layout(ws)
        if layout:
            return max(0, layout["data_start"] - (layout["header_rows"][-1] + 1))
        # legacy fallback
        row = START_ROW
        while any(ws.cell(row=row, column=c).value not in (None, "") for c in (2, 3)):
            row += 1
        return max(0, row - START_ROW)
    except Exception:
        return 0


def export_to_excel(records: list, template_bytes: bytes, append: bool = True) -> bytes:
    """Fill the template by auto-detecting its columns from the header row(s).

    append=True  → add after the last existing record (keeps prior data).
    append=False → start at the first data row (overwrite).
    """
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        f.write(template_bytes)
        tmp = f.name
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            wb = load_workbook(tmp)
        ws = _pick_sheet(wb)

        layout = _detect_layout(ws)
        if layout and layout["mappings"]:
            mappings = layout["mappings"]
            first_data = layout["header_rows"][-1] + 1
            start = layout["data_start"] if append else first_data
        else:                                           # legacy fixed layout
            mappings = [(c, k, "") for c, k in COLUMN_MAP]
            first_data = START_ROW
            row = START_ROW
            while any(ws.cell(row=row, column=cc).value not in (None, "") for cc in (2, 3)):
                row += 1
            start = row if append else START_ROW

        for i, rec in enumerate(records):
            row = start + i
            for col, field, header_text in mappings:
                value = _value_for(field, rec, header_text)
                cell = ws.cell(row=row, column=col, value=value)
                if field in TEXT_FIELDS:                # no 1.31E+13 on long digits
                    cell.number_format = "@"
                if field == "arabic_name" or (value and ARABIC_RE.search(str(value))):
                    _right_align(cell)

        _reattach_dropdowns(wb, ws)

        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()
    finally:
        os.unlink(tmp)


def _excel_text_guard(value) -> str:
    s = "" if value is None else str(value)
    return f'="{s}"' if s.isdigit() and len(s) >= 12 else s


def export_to_csv(records: list) -> bytes:
    """All records as UTF-8 (BOM) CSV bytes — Excel-safe Arabic + text numbers."""
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=CSV_FIELDS, extrasaction="ignore")
    writer.writeheader()
    for rec in records:
        row = dict(rec)
        row["account_number"] = _excel_text_guard(row.get("account_number", ""))
        writer.writerow(row)
    return buf.getvalue().encode("utf-8-sig")
